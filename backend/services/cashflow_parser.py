"""Parser for PMS cash flow Excel files.

Files are simple flat ledgers (no embedded client headers like NAV/transaction files).
Each data row is one cash flow event: either a Receipt (inflow) or Payment (outflow).

Columns:
  A: Date (DD-MMM-YYYY string, e.g. "19-Sep-2020")
  B: Branch (always "HO")
  C: UCC (client code with trailing spaces)
  D: Account Head ("CLIENT NAME [UCC]")
  E: Receipts (money IN — capital added by client)
  F: Payments (money OUT — withdrawals + PMS fees)
  G: Balance (running cumulative — ignore)
  H: (empty/"Dr." — ignore)
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Column indices (0-based)
_COL_DATE = 0
_COL_UCC = 2
_COL_ACCOUNT_HEAD = 3
_COL_RECEIPTS = 4
_COL_PAYMENTS = 5


def parse_cashflow_file(filepath: str | Path) -> list[dict]:
    """
    Parse a single cash flow Excel file.

    Returns list of dicts with keys:
        client_code, client_name, date, flow_type, amount
    """
    filepath = Path(filepath)
    logger.info("Parsing cash flow file: %s", filepath.name)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    records: list[dict] = []
    header_skipped = False

    for row_cells in ws.iter_rows(values_only=True):
        vals = list(row_cells)

        # Skip header row
        if not header_skipped:
            header_skipped = True
            continue

        # Skip rows where Date is None or empty (Opening Balance, subtotals, grand totals)
        date_cell = vals[_COL_DATE] if len(vals) > _COL_DATE else None
        if date_cell is None:
            continue
        date_str = str(date_cell).strip()
        if not date_str or date_str.startswith("GRAND") or date_str.startswith("Opening"):
            continue

        # Parse date
        try:
            if isinstance(date_cell, datetime):
                flow_date = date_cell.date()
            else:
                flow_date = datetime.strptime(date_str, "%d-%b-%Y").date()
        except (ValueError, TypeError):
            continue  # Not a data row

        # Extract UCC
        ucc_raw = vals[_COL_UCC] if len(vals) > _COL_UCC else None
        if ucc_raw is None:
            continue
        client_code = str(ucc_raw).strip()
        if not client_code:
            continue

        # Extract client name from Account Head: "CLIENT NAME [UCC]"
        account_head = str(vals[_COL_ACCOUNT_HEAD]).strip() if len(vals) > _COL_ACCOUNT_HEAD and vals[_COL_ACCOUNT_HEAD] else ""
        name_match = re.match(r"^(.+?)\s*\[", account_head)
        client_name = name_match.group(1).strip() if name_match else account_head

        # Extract amounts as Decimal (financial data — never float)
        receipts = _safe_decimal(vals[_COL_RECEIPTS] if len(vals) > _COL_RECEIPTS else None)
        payments = _safe_decimal(vals[_COL_PAYMENTS] if len(vals) > _COL_PAYMENTS else None)

        if receipts > 0:
            records.append({
                "client_code": client_code,
                "client_name": client_name,
                "date": flow_date,
                "flow_type": "INFLOW",
                "amount": receipts,
                "description": f"Capital inflow - {filepath.stem}",
            })

        if payments > 0:
            records.append({
                "client_code": client_code,
                "client_name": client_name,
                "date": flow_date,
                "flow_type": "OUTFLOW",
                "amount": payments,
                "description": f"Capital outflow - {filepath.stem}",
            })

    wb.close()
    logger.info("Parsed %d cash flow records from %s", len(records), filepath.name)
    return records


def parse_all_cashflow_files(directory: str | Path) -> list[dict]:
    """Parse all 'Cash outflow and inflow-*.xlsx' files in directory."""
    directory = Path(directory)
    files = sorted(directory.glob("Cash outflow and inflow-*.xlsx"))
    if not files:
        logger.warning("No cash flow files found in %s", directory)
        return []

    all_records: list[dict] = []
    for f in files:
        records = parse_cashflow_file(f)
        all_records.extend(records)

    logger.info("Total cash flow records across %d files: %d", len(files), len(all_records))
    return all_records


def _safe_decimal(val) -> Decimal:
    """Safely convert a cell value to Decimal, defaulting to 0."""
    if val is None:
        return Decimal("0")
    try:
        d = Decimal(str(val))
        if d.is_nan():
            return Decimal("0")
        return d
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")
