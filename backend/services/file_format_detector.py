"""Sniff a PMS backoffice .xlsx to detect what kind of file it is.

Used at upload time to reject files dropped into the wrong slot — e.g. a
Transaction file uploaded into the Holdings slot — before they can overwrite
clean data via the UPSERT-on-natural-key parsers.

Detection is strictly header-based. We scan the first ~10 rows, normalize the
cells, and score each row against a fingerprint of required tokens per known
format. The best scoring row/format combo wins if confidence ≥ threshold.

Known formats:
  - NAV            → cpp_nav_series loader
  - TRANSACTIONS   → cpp_transactions loader
  - HOLDINGS       → cpp_holdings loader (equity OR ETF — same layout, same parser)
  - CASHFLOWS     → cpp_cashflows loader

The HOLDINGS fingerprint can't be split into EQUITY vs ETF from headers alone —
that's a content distinction, not a structural one. Both holdings slots accept
the same fingerprint; the caller chooses which bucket to ingest into.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ── Public types ────────────────────────────────────────────────────────────
FileFormat = Literal["NAV", "TRANSACTIONS", "HOLDINGS", "CASHFLOWS"]
UploadSlot = Literal["NAV", "TRANSACTIONS", "EQUITY_HOLDINGS", "ETF_HOLDINGS", "CASHFLOWS"]


@dataclass(frozen=True)
class DetectionResult:
    detected: FileFormat | None
    confidence: float  # 0.0 – 1.0
    header_row_index: int | None  # 0-based row where the header was found
    header_preview: tuple[str, ...]  # first ~12 non-empty cells of that row
    reason: str


class FileFormatMismatch(Exception):
    """Raised by ``assert_format`` when the upload slot doesn't accept the detected format.

    The ``detail`` attribute is a user-facing message suitable for an HTTP 400
    response. ``slot`` and ``result`` are kept for logging / telemetry.
    """

    def __init__(self, slot: UploadSlot, result: DetectionResult, detail: str) -> None:
        self.slot = slot
        self.result = result
        self.detail = detail
        super().__init__(detail)


# ── Fingerprints ────────────────────────────────────────────────────────────
# For each format: (required_tokens, threshold)
#   required_tokens = set of lowercase substring tokens the header cells must
#                     collectively cover. A token "matches" if it appears as a
#                     substring of any normalized cell in the row.
#   threshold       = min fraction of tokens that must match to count as a hit.
#
# Tokens are chosen to be DISCRIMINATING — "corpus" only appears in NAV,
# "stno" only in TXN, "market rate" only in holdings, "receipt"/"payment" in
# cashflows. Generic tokens like "ucc" and "date" appear in many files, so we
# stack them with more-distinctive ones to avoid false positives.
_FINGERPRINTS: tuple[tuple[FileFormat, frozenset[str], float], ...] = (
    ("CASHFLOWS", frozenset({"receipt", "payment", "account head"}), 0.66),
    ("NAV", frozenset({"ucc", "date", "nav", "corpus"}), 0.75),
    ("TRANSACTIONS", frozenset({"ucc", "script", "stno", "exch"}), 0.75),
    ("HOLDINGS", frozenset({"ucc", "share", "isin", "market rate", "% holding"}), 0.60),
)

# Which detected format each upload-slot accepts.
_SLOT_ACCEPTS: dict[UploadSlot, frozenset[FileFormat]] = {
    "NAV": frozenset({"NAV"}),
    "TRANSACTIONS": frozenset({"TRANSACTIONS"}),
    "EQUITY_HOLDINGS": frozenset({"HOLDINGS"}),
    "ETF_HOLDINGS": frozenset({"HOLDINGS"}),
    "CASHFLOWS": frozenset({"CASHFLOWS"}),
}

# Human-readable labels for error messages.
_SLOT_LABELS: dict[UploadSlot, str] = {
    "NAV": "NAV Report",
    "TRANSACTIONS": "Transaction Report",
    "EQUITY_HOLDINGS": "Equity Holding Report",
    "ETF_HOLDINGS": "ETF Holding Report",
    "CASHFLOWS": "Cash Flow Report",
}

_FORMAT_LABELS: dict[FileFormat, str] = {
    "NAV": "NAV Report",
    "TRANSACTIONS": "Transaction Report",
    "HOLDINGS": "Holding Report",
    "CASHFLOWS": "Cash Flow Report",
}

_EXPECTED_COLUMNS: dict[UploadSlot, str] = {
    "NAV": "UCC, Date, Corpus, NAV, Liquidity %, High Water Mark",
    "TRANSACTIONS": "UCC, Script, Exch, Stno, Buy Qty, Sale Qty, …",
    "EQUITY_HOLDINGS": "UCC, Share (PMS), ISIN, Stock (qty), Market Rate, % Holding Market",
    "ETF_HOLDINGS": "UCC, Share (PMS), ISIN, Stock (qty), Market Rate, % Holding Market",
    "CASHFLOWS": "Date, Branch, UCC, Account Head, Receipts, Payments",
}

# How many rows from the top of the sheet to scan when looking for the header.
_MAX_HEADER_SCAN_ROWS = 10


# ── Helpers ─────────────────────────────────────────────────────────────────
def _normalize_cell(value: object) -> str:
    """Lowercase + collapse whitespace + strip Excel escape sequences."""
    if value is None:
        return ""
    return " ".join(
        str(value)
        .lower()
        .replace("_x000d_", " ")
        .replace("\r\n", " ")
        .replace("\n", " ")
        .split()
    )


def _scan_header_rows(filepath: Path) -> list[list[str]]:
    """Read up to _MAX_HEADER_SCAN_ROWS from the primary sheet.

    Prefers 'Sheet1' (the actual data sheet in newer NAV exports) over
    ``wb.active`` — some backoffice exports make a summary sheet active.
    Returns a list of normalized rows. Each row is a list of lowercase strings.
    """
    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        if ws is None:
            return []
        rows: list[list[str]] = []
        for count, raw_row in enumerate(ws.iter_rows(values_only=True)):
            if count >= _MAX_HEADER_SCAN_ROWS:
                break
            rows.append([_normalize_cell(c) for c in raw_row])
        return rows
    finally:
        wb.close()


def _score_row(row: list[str], required: frozenset[str]) -> float:
    """Fraction of required tokens found as substrings anywhere in the row."""
    if not required:
        return 0.0
    non_empty = [c for c in row if c]
    if not non_empty:
        return 0.0
    matched = sum(1 for tok in required if any(tok in cell for cell in non_empty))
    return matched / len(required)


def _row_preview(raw_row: list[str], limit: int = 12) -> tuple[str, ...]:
    """First ``limit`` non-empty cells from a scanned row, preserving order."""
    return tuple(c for c in raw_row if c)[:limit]


# ── Public API ──────────────────────────────────────────────────────────────
def detect_file_format(filepath: str | Path) -> DetectionResult:
    """Inspect the header rows of an .xlsx and classify it.

    Returns a DetectionResult — ``detected`` is None when no fingerprint
    scored above its threshold. ``confidence`` is the best-match score.

    Does NOT validate data rows. Intended for cheap pre-upload checks.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    rows = _scan_header_rows(path)
    if not rows:
        return DetectionResult(
            detected=None,
            confidence=0.0,
            header_row_index=None,
            header_preview=(),
            reason="Workbook appears empty — no rows readable.",
        )

    best: tuple[FileFormat, float, int, list[str]] | None = None
    for row_idx, row in enumerate(rows):
        for fmt, required, threshold in _FINGERPRINTS:
            score = _score_row(row, required)
            if score < threshold:
                continue
            if best is None or score > best[1]:
                best = (fmt, score, row_idx, row)

    if best is None:
        first_preview = _row_preview(rows[0])
        return DetectionResult(
            detected=None,
            confidence=0.0,
            header_row_index=None,
            header_preview=first_preview,
            reason=(
                "No recognized header pattern in the first "
                f"{len(rows)} rows. First row cells: {list(first_preview)}"
            ),
        )

    fmt, score, row_idx, row = best
    return DetectionResult(
        detected=fmt,
        confidence=score,
        header_row_index=row_idx,
        header_preview=_row_preview(row),
        reason=f"Matched {fmt} with confidence {score:.0%} at row {row_idx + 1}",
    )


def assert_format(filepath: str | Path, slot: UploadSlot) -> DetectionResult:
    """Detect the file's format; raise ``FileFormatMismatch`` if the slot rejects it.

    Call this synchronously at the upload endpoint, BEFORE saving to temp
    and dispatching a background ingestion task. That way a wrong-slot
    upload produces a 400 error the admin sees immediately, and the
    upsert-on-natural-key parsers never run against the wrong shape.
    """
    result = detect_file_format(filepath)
    accepted = _SLOT_ACCEPTS.get(slot, frozenset())
    if result.detected is not None and result.detected in accepted:
        return result

    slot_label = _SLOT_LABELS.get(slot, slot)
    expected_cols = _EXPECTED_COLUMNS.get(slot, "")
    header_cells = ", ".join(result.header_preview[:8]) or "(none found)"

    if result.detected is None:
        detail = (
            f"This file doesn't match the {slot_label} format. "
            f"We scanned the first {_MAX_HEADER_SCAN_ROWS} rows and couldn't recognize the column layout. "
            f"Expected columns for {slot_label}: {expected_cols}. "
            f"First row we saw: {header_cells}."
        )
    else:
        detected_label = _FORMAT_LABELS[result.detected]
        detail = (
            f"This looks like a {detected_label} but you uploaded it into the {slot_label} slot. "
            f"Please use the {detected_label} upload option, or choose a {slot_label} file for this slot. "
            f"Expected columns here: {expected_cols}. "
            f"Detected columns: {header_cells}."
        )

    logger.warning(
        "File format mismatch: slot=%s detected=%s confidence=%.2f preview=%r",
        slot,
        result.detected,
        result.confidence,
        result.header_preview,
    )
    raise FileFormatMismatch(slot=slot, result=result, detail=detail)
