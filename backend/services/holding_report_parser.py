"""Stateful .xlsx parser for PMS Holding Report files.

Uses openpyxl read_only mode for memory-efficient parsing.

Supports two column layouts dynamically detected from the header row:

  16-col (full firm report — has Family Group):
    0: UCC, 1: Family Group, 2: Share (PMS), 3: ISIN, 4: Stock (qty),
    5: Cost (Rs.), 6: Total Cost, 7: % Holding Cost, 8: % Cumul,
    9: Market Rate, 10: Market Rate Date, 11: Market Value (Rs.),
    12: Notional P/L, 13: ROI [%], 14: % Holding Market, 15: % Cumul

  15-col (single-client report — no Family Group):
    0: UCC, 1: Share (PMS), 2: ISIN, 3: Stock (qty),
    4: Cost (Rs.), 5: Total Cost, 6: % Holding Cost, 7: % Cumul,
    8: Market Rate, 9: Market Rate Date, 10: Market Value (Rs.),
    11: Notional P/L, 12: ROI [%], 13: % Holding Market, 14: % Cumul
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ── Backward-compatible default indices (16-col / Family Group format) ──────
_COL_UCC = 0
_COL_FAMILY_GROUP = 1
_COL_SHARE = 2
_COL_ISIN = 3
_COL_QTY = 4
_COL_AVG_COST = 5
_COL_TOTAL_COST = 6
_COL_HOLDING_COST_PCT = 7
_COL_MARKET_PRICE = 9
_COL_MARKET_DATE = 10
_COL_MARKET_VALUE = 11
_COL_NOTIONAL_PNL = 12
_COL_ROI_PCT = 13
_COL_HOLDING_MARKET_PCT = 14

_MIN_COLUMNS = 14  # At least 14 cols needed (15-col format minimum usable)

# Date format used in Market Rate Date column
_MARKET_DATE_FORMAT = "%d/%m/%Y"


def _safe_decimal(value: Any) -> Decimal | None:
    """Convert a cell value to Decimal.

    Returns None for None/empty values so callers can distinguish
    missing data from zero — critical for financial integrity.
    """
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw.lower() in ("nan", "none", "-", ""):
        return None
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        logger.debug("Non-numeric cell value: %r", raw)
        return None


def _parse_market_date(value: Any) -> date | None:
    """Parse the Market Rate Date cell.

    Accepts:
      - Python date/datetime objects (openpyxl may return these)
      - String in DD/MM/YYYY format
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw or raw.lower() in ("nan", "none"):
        return None
    try:
        return datetime.strptime(raw, _MARKET_DATE_FORMAT).date()
    except ValueError:
        logger.debug("Unparseable market date: %r", raw)
        return None


def _parse_share(raw: str) -> tuple[str, str]:
    """Extract symbol and instrument type from a Share (PMS) cell.

    Example: "CPSEETF EQ                    " → ("CPSEETF", "EQ")
    Example: "GLENMARK EQ                   " → ("GLENMARK", "EQ")
    Example: "LIQUIDCASE EQ                 " → ("LIQUIDCASE", "EQ")

    Returns:
        (symbol, instrument_type) — both uppercase.
        instrument_type is empty string if not present.
    """
    stripped = raw.strip()
    parts = stripped.split()
    if not parts:
        return ("", "")
    symbol = parts[0].upper()
    instrument_type = parts[1].upper() if len(parts) > 1 else ""
    return symbol, instrument_type


def _build_col_map(header_row: tuple) -> dict[str, int]:
    """Build a column index map from the header row.

    Handles both 15-col (no Family Group) and 16-col (with Family Group)
    formats by detecting column positions from header keywords.

    Returns a dict mapping internal field names to column indices.
    Returns empty dict if the row doesn't look like a valid header.
    """
    col_map: dict[str, int] = {}

    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        # Normalize: lowercase, collapse whitespace and Excel escape sequences
        norm = (
            str(cell)
            .lower()
            .replace("_x000d_", " ")
            .replace("\r\n", " ")
            .replace("\n", " ")
            .strip()
        )
        norm = " ".join(norm.split())  # collapse multiple spaces

        if norm == "ucc":
            col_map["ucc"] = idx
        elif "family" in norm:
            col_map["family_group"] = idx
        elif "share" in norm or "(pms)" in norm:
            # "Share (PMS)" — the symbol column
            col_map["share"] = idx
        elif norm == "isin":
            col_map["isin"] = idx
        elif "stock" in norm:
            # "Stock (qty)" or just "Stock" — quantity column
            col_map["qty"] = idx
        elif "total" in norm and "cost" in norm:
            # "Total Cost" — must check before bare "cost"
            col_map["total_cost"] = idx
        elif "cost" in norm and "%" not in norm:
            # "Cost (Rs.)" — average cost per share
            col_map.setdefault("avg_cost", idx)
        elif "market" in norm and "rate" in norm and "date" in norm:
            # "Market Rate Date" — must check before bare "Market Rate"
            col_map["market_date"] = idx
        elif "market" in norm and "value" in norm:
            # "Market Value (Rs.)"
            col_map["market_value"] = idx
        elif "market" in norm and "rate" in norm:
            # "Market Rate"
            col_map.setdefault("market_price", idx)
        elif "notional" in norm or ("p/l" in norm and "roi" not in norm):
            col_map["notional_pnl"] = idx
        elif "roi" in norm:
            col_map["roi_pct"] = idx
        elif "% holding" in norm and "market" in norm and "cumul" not in norm:
            col_map["holding_market_pct"] = idx
        elif "% holding" in norm and "cost" in norm and "cumul" not in norm:
            col_map["holding_cost_pct"] = idx

    return col_map


def _is_data_row(row: tuple, col_map: dict[str, int] | None = None) -> bool:
    """Return True if this row looks like a holding data row.

    A valid data row has:
      - A non-empty UCC (col 0)
      - A non-empty Share/symbol column
      - A numeric quantity column
    Subtotal, header, and empty rows fail these checks.

    Args:
        row: The row tuple from the workbook.
        col_map: Optional column map from _build_col_map(). If None, uses
                 default 16-col (Family Group) indices for backward compatibility.
    """
    if col_map is not None:
        col_ucc = col_map.get("ucc", _COL_UCC)
        col_share = col_map.get("share", _COL_SHARE)
        col_qty = col_map.get("qty", _COL_QTY)
    else:
        col_ucc, col_share, col_qty = _COL_UCC, _COL_SHARE, _COL_QTY

    min_len = max(col_ucc, col_share, col_qty) + 1
    if len(row) < min_len:
        return False

    ucc_raw = row[col_ucc]
    share_raw = row[col_share]
    qty_raw = row[col_qty]

    if ucc_raw is None or share_raw is None or qty_raw is None:
        return False

    ucc = str(ucc_raw).strip()
    share = str(share_raw).strip()
    if not ucc or not share:
        return False

    # Quantity must be numeric
    try:
        float(str(qty_raw).replace(",", ""))
    except ValueError:
        return False

    return True


def parse_holding_report(filepath: str | Path) -> list[dict]:
    """Parse a PMS Holding Report .xlsx file.

    Dynamically detects whether the file has 15 cols (no Family Group) or
    16 cols (with Family Group) by reading the header row and mapping
    column positions from keywords.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        List of dicts, one per holding row, with keys:
            ucc, family_group, symbol, share_raw, isin, instrument_type,
            quantity, avg_cost, total_cost, holding_cost_pct,
            market_price, market_date, market_value,
            notional_pnl, roi_pct, holding_market_pct

        Financial fields are Decimal | None.
        market_date is date | None.
        family_group is empty string when the column is absent.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the workbook has no active sheet.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Holding report file not found: {filepath}")

    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Holding report workbook has no active sheet")

    records: list[dict] = []
    row_count = 0
    skipped_count = 0
    header_skipped = False
    ucc_row_counts: dict[str, int] = {}

    # Column map is built from the first (header) row
    col_map: dict[str, int] = {}

    for row in ws.iter_rows(values_only=True):
        row_count += 1

        # ── First row: detect column layout from header ─────────────────────
        if not header_skipped:
            header_skipped = True
            detected = _build_col_map(row)
            if detected and "ucc" in detected and "share" in detected:
                col_map = detected
                logger.info(
                    "Holding parser: detected column map — %s",
                    {k: v for k, v in col_map.items()},
                )
            else:
                # Fallback: use default 16-col (Family Group) indices
                col_map = {
                    "ucc": _COL_UCC,
                    "family_group": _COL_FAMILY_GROUP,
                    "share": _COL_SHARE,
                    "isin": _COL_ISIN,
                    "qty": _COL_QTY,
                    "avg_cost": _COL_AVG_COST,
                    "total_cost": _COL_TOTAL_COST,
                    "holding_cost_pct": _COL_HOLDING_COST_PCT,
                    "market_price": _COL_MARKET_PRICE,
                    "market_date": _COL_MARKET_DATE,
                    "market_value": _COL_MARKET_VALUE,
                    "notional_pnl": _COL_NOTIONAL_PNL,
                    "roi_pct": _COL_ROI_PCT,
                    "holding_market_pct": _COL_HOLDING_MARKET_PCT,
                }
                logger.warning(
                    "Holding parser: header detection failed, using default 16-col map"
                )
            continue

        if not _is_data_row(row, col_map):
            skipped_count += 1
            continue

        col_ucc = col_map.get("ucc", _COL_UCC)
        col_family = col_map.get("family_group")
        col_share = col_map.get("share", _COL_SHARE)
        col_isin = col_map.get("isin", _COL_ISIN)
        col_qty = col_map.get("qty", _COL_QTY)
        col_avg_cost = col_map.get("avg_cost", _COL_AVG_COST)
        col_total_cost = col_map.get("total_cost", _COL_TOTAL_COST)
        col_holding_cost_pct = col_map.get("holding_cost_pct", _COL_HOLDING_COST_PCT)
        col_market_price = col_map.get("market_price", _COL_MARKET_PRICE)
        col_market_date = col_map.get("market_date", _COL_MARKET_DATE)
        col_market_value = col_map.get("market_value", _COL_MARKET_VALUE)
        col_notional_pnl = col_map.get("notional_pnl", _COL_NOTIONAL_PNL)
        col_roi_pct = col_map.get("roi_pct", _COL_ROI_PCT)
        col_holding_market_pct = col_map.get("holding_market_pct", _COL_HOLDING_MARKET_PCT)

        ucc = str(row[col_ucc]).strip()

        # Family group: empty string when the column is absent
        family_group = ""
        if col_family is not None and col_family < len(row) and row[col_family] is not None:
            family_group = str(row[col_family]).strip()

        share_raw_cell = str(row[col_share]).strip() if row[col_share] is not None else ""
        isin_raw = row[col_isin] if col_isin < len(row) else None
        isin = str(isin_raw).strip() if isin_raw is not None else ""

        symbol, instrument_type = _parse_share(share_raw_cell)
        if not symbol:
            logger.debug("Row %d: empty symbol after parse, skipping", row_count)
            skipped_count += 1
            continue

        def _get(col_idx: int | None, default_idx: int) -> Any:
            idx = col_idx if col_idx is not None else default_idx
            return row[idx] if idx < len(row) else None

        market_date = _parse_market_date(_get(col_market_date, _COL_MARKET_DATE))
        quantity = _safe_decimal(_get(col_qty, _COL_QTY))
        avg_cost = _safe_decimal(_get(col_avg_cost, _COL_AVG_COST))
        total_cost = _safe_decimal(_get(col_total_cost, _COL_TOTAL_COST))
        holding_cost_pct = _safe_decimal(_get(col_holding_cost_pct, _COL_HOLDING_COST_PCT))
        market_price = _safe_decimal(_get(col_market_price, _COL_MARKET_PRICE))
        market_value = _safe_decimal(_get(col_market_value, _COL_MARKET_VALUE))
        notional_pnl = _safe_decimal(_get(col_notional_pnl, _COL_NOTIONAL_PNL))
        roi_pct = _safe_decimal(_get(col_roi_pct, _COL_ROI_PCT))
        holding_market_pct = _safe_decimal(_get(col_holding_market_pct, _COL_HOLDING_MARKET_PCT))

        records.append(
            {
                "ucc": ucc,
                "family_group": family_group,
                "symbol": symbol,
                "share_raw": share_raw_cell,
                "isin": isin,
                "instrument_type": instrument_type,
                "quantity": quantity,
                "avg_cost": avg_cost,
                "total_cost": total_cost,
                "holding_cost_pct": holding_cost_pct,
                "market_price": market_price,
                "market_date": market_date,
                "market_value": market_value,
                "notional_pnl": notional_pnl,
                "roi_pct": roi_pct,
                "holding_market_pct": holding_market_pct,
            }
        )

        ucc_row_counts[ucc] = ucc_row_counts.get(ucc, 0) + 1

    wb.close()

    for ucc, count in sorted(ucc_row_counts.items()):
        logger.info("Holding parser: UCC %s — %d holding rows", ucc, count)

    logger.info(
        "Holding parser complete: %d records from %d UCCs (%d rows scanned, %d skipped)",
        len(records),
        len(ucc_row_counts),
        row_count,
        skipped_count,
    )
    return records


def holding_report_summary(records: list[dict]) -> dict:
    """Return summary statistics for a parsed holding report.

    Args:
        records: Output of parse_holding_report().

    Returns:
        Dict with keys:
            total_rows:      int — number of holding records
            unique_uccs:     int — number of distinct client codes
            unique_symbols:  int — number of distinct stock symbols
            market_date:     date | None — most common market date across all rows
            uccs:            list[str] — sorted list of UCC codes
    """
    if not records:
        return {
            "total_rows": 0,
            "unique_uccs": 0,
            "unique_symbols": 0,
            "market_date": None,
            "uccs": [],
        }

    unique_uccs: set[str] = set()
    unique_symbols: set[str] = set()
    date_counts: dict[date, int] = {}

    for rec in records:
        unique_uccs.add(rec["ucc"])
        unique_symbols.add(rec["symbol"])
        mdate = rec.get("market_date")
        if mdate is not None:
            date_counts[mdate] = date_counts.get(mdate, 0) + 1

    # Most common market date (typically all rows share the same date)
    market_date: date | None = (
        max(date_counts, key=lambda d: date_counts[d]) if date_counts else None
    )

    return {
        "total_rows": len(records),
        "unique_uccs": len(unique_uccs),
        "unique_symbols": len(unique_symbols),
        "market_date": market_date,
        "uccs": sorted(unique_uccs),
    }
