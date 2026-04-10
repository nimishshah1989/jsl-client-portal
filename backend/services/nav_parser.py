"""Stateful .xlsx parser for PMS NAV Report files.

Uses openpyxl read_only mode for memory-efficient parsing of 35MB files.
See FILE_FORMAT_SPEC.md for full format documentation.

Row types:
  1. Client name header — "FULL NAME [CODE]" in UCC column
  2. Data row — UCC matches current client code, has valid Date
  3. Grand total / subtotal — UCC is None, skip
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Column indices after header row (0-based)
# Actual file has 10 columns:
# UCC, Date, Corpus, Equity Holding At Mkt, Investments in ETF,
# Cash And Cash Equivalent, Bank Balance, NAV, Liquidity %, High Water Mark
_COL_UCC = 0
_COL_DATE = 1
_COL_CORPUS = 2
_COL_EQUITY = 3
_COL_ETF = 4        # "Investments in ETF" — not in original spec
_COL_CASH = 5
_COL_BANK = 6
_COL_NAV = 7
_COL_LIQUIDITY = 8
_COL_HWM = 9

# Regex for client name header: "FULL NAME [CODE]"
_NAME_PATTERN = re.compile(r"^(.+?)\s*\[(\w+)\]$")

# Date formats in NAV file — backoffice exports use either format
_DATE_FORMATS = [
    "%d-%b-%Y",    # DD-MMM-YYYY e.g. "28-Sep-2020"
    "%d/%m/%Y",    # DD/MM/YYYY  e.g. "26/03/2026"
]


def _safe_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, returning Decimal('0') for non-numeric."""
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_nav_date(value: object) -> datetime | None:
    """Parse a NAV date cell — handles DD-MMM-YYYY, DD/MM/YYYY, and datetime objects."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    raw = str(value).strip()
    if not raw or raw.lower() == "nan":
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    logger.debug("Unparseable NAV date: %r", raw)
    return None


# Mapping from normalized header keywords to our internal field names
_HEADER_KEYWORDS: dict[str, list[str]] = {
    "ucc": ["ucc"],
    "date": ["date"],
    "client": ["client"],
    "codes": ["codes", "code"],
    "corpus": ["corpus"],
    "equity": ["equity holding", "equity_holding", "equity"],
    "etf": ["investments in etf", "investments_in_etf", "etf"],
    "cash": ["cash and cash equivalent", "cash_and_cash", "cash and_cash"],
    "bank": ["bank balance", "bank_balance", "bank"],
    "nav": ["nav"],
    "liquidity": ["liquidity %", "liquidity_pct", "liquidity"],
    "hwm": ["high water mark", "high_water_mark", "water mark"],
}


def _build_column_map(header_row: tuple) -> dict[str, int] | None:
    """Build a column index map from the header row.

    Matches header cells to internal field names using keyword matching.
    Returns None if the row does not look like a header (no 'ucc' or 'nav' found).
    """
    if header_row is None:
        return None

    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        # Normalize: lowercase, strip whitespace/newlines
        norm = str(cell).lower().replace("\r\n", " ").replace("\n", " ").replace("_x000d_", " ").strip()
        norm = " ".join(norm.split())  # collapse multiple spaces

        for field_name, keywords in _HEADER_KEYWORDS.items():
            if field_name in col_map:
                continue  # already matched
            for kw in keywords:
                if kw in norm:
                    col_map[field_name] = idx
                    break

    # Must have at least UCC and NAV to be a valid header
    if "ucc" not in col_map or "nav" not in col_map:
        return None
    # Must have date
    if "date" not in col_map:
        return None

    return col_map


def parse_nav_file(filepath: str | Path) -> list[dict]:
    """
    Parse a PMS NAV Report .xlsx file.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        List of dicts with keys:
            client_code, client_name, date, corpus, equity_value,
            cash_value, bank_balance, nav, liquidity_pct, high_water_mark
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"NAV file not found: {filepath}")

    wb = load_workbook(str(filepath), read_only=True, data_only=True)

    # Prefer Sheet1 (the actual NAV data) over the active sheet, which may
    # be a pivot/summary sheet in newer backoffice exports.
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    if ws is None:
        wb.close()
        raise ValueError("NAV workbook has no active sheet")

    records: list[dict] = []
    current_client_code: str | None = None
    current_client_name: str | None = None
    row_count = 0
    col_map: dict[str, int] | None = None
    clients_seen: set[str] = set()

    for row in ws.iter_rows(values_only=True):
        row_count += 1

        # --- Detect header row and build column map dynamically ---
        if col_map is None:
            col_map = _build_column_map(row)
            if col_map is not None:
                logger.info("NAV parser: column map built — %s", col_map)
            continue

        if len(row) < 6:
            continue

        ucc_raw = row[col_map["ucc"]]

        # Subtotal / grand total rows have None UCC — skip
        if ucc_raw is None:
            continue

        ucc = str(ucc_raw).strip()
        if not ucc or ucc.lower() == "nan":
            continue

        # Check for client name header: "FULL NAME [CODE]"
        name_match = _NAME_PATTERN.match(ucc)
        if name_match:
            current_client_name = name_match.group(1).strip()
            current_client_code = name_match.group(2).strip()
            if current_client_code not in clients_seen:
                clients_seen.add(current_client_code)
                logger.info(
                    "NAV parser: processing client %s (%s) — client #%d",
                    current_client_code,
                    current_client_name,
                    len(clients_seen),
                )
            continue

        # Flat format: no client headers — UCC + optional Codes column
        if current_client_code is None and "codes" in col_map:
            codes_val = row[col_map["codes"]]
            if codes_val is not None:
                current_client_code = str(codes_val).strip()
                # Use Client column or UCC for name
                if "client" in col_map and row[col_map["client"]]:
                    current_client_name = str(row[col_map["client"]]).strip()
                else:
                    current_client_name = ucc.rstrip()

        # Flat format: each row has UCC directly, no name-header grouping
        if current_client_code is None or ucc.rstrip() != current_client_code:
            # Try treating this row as a self-contained data row
            if "codes" in col_map:
                codes_val = row[col_map["codes"]]
                if codes_val is not None:
                    current_client_code = str(codes_val).strip()
                    if "client" in col_map and row[col_map["client"]]:
                        current_client_name = str(row[col_map["client"]]).strip()
                    else:
                        current_client_name = ucc.rstrip()
                else:
                    continue
            else:
                continue

        nav_date = _parse_nav_date(row[col_map["date"]])
        if nav_date is None:
            continue

        nav_val = _safe_decimal(row[col_map["nav"]])
        if nav_val == 0:
            logger.debug(
                "Skipping zero-NAV row for %s on %s",
                current_client_code,
                nav_date.date(),
            )
            continue

        records.append(
            {
                "client_code": current_client_code,
                "client_name": current_client_name,
                "date": nav_date,
                "corpus": _safe_decimal(row[col_map["corpus"]]) if "corpus" in col_map else Decimal("0"),
                "equity_value": _safe_decimal(row[col_map["equity"]]) if "equity" in col_map else Decimal("0"),
                "etf_value": _safe_decimal(row[col_map["etf"]]) if "etf" in col_map else Decimal("0"),
                "cash_value": _safe_decimal(row[col_map["cash"]]) if "cash" in col_map else Decimal("0"),
                "bank_balance": _safe_decimal(row[col_map["bank"]]) if "bank" in col_map else Decimal("0"),
                "nav": nav_val,
                "liquidity_pct": _safe_decimal(row[col_map["liquidity"]]) if "liquidity" in col_map else Decimal("0"),
                "high_water_mark": _safe_decimal(row[col_map["hwm"]]) if "hwm" in col_map else Decimal("0"),
            }
        )

        # Reset for next row in flat format (each row is independent)
        if "codes" in col_map:
            current_client_code = None
            current_client_name = None

    wb.close()
    logger.info(
        "NAV parser complete: %d records from %d clients (%d rows scanned)",
        len(records),
        len(clients_seen),
        row_count,
    )
    return records
