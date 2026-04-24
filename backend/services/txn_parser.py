"""Stateful .xlsx parser for PMS Transaction Report files.

Uses openpyxl read_only mode for memory-efficient parsing of 35MB files.
See FILE_FORMAT_SPEC.md for full format documentation.

Column layout is AUTO-DETECTED from the header row — no hardcoded indices.

Two formats are supported:

  Old (20-col):  UCC | Script | Exch | Stno | Buy×8 | Sale×8
  New (21-col):  UCC | ISIN | Script | Exch | Stno | Buy×8 | Sale×8

The Buy and Sale blocks are always 8 columns wide:
  +0 Quantity, +1 Net Rate, +2 GST, +3 Other Charges, +4 STT,
  +5 Cost Rate, +6 Amount With Cost, +7 Amount Without STT

Row types:
  0. Sub-header row 0 — format detection, then skip
  1. Sub-header row 1 — skip
  2. Client name header — "FULL NAME [CODE]"
  3. Date separator — "     Date :DD/MM/YY"
  4. Transaction data — UCC = client code, Script present
  5. Daily subtotal — UCC is None, skip
  6. Grand total — UCC is None, skip
"""

from __future__ import annotations

import logging
import re
from collections import namedtuple
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Regex patterns
_NAME_PATTERN = re.compile(r"^(.+?)\s*\[(\w+)\]$")
_DATE_PATTERN = re.compile(r"Date\s*:\s*(\d{2}/\d{2}/\d{2})")

# Known instrument type suffixes — last token is instrument type when it matches
_KNOWN_INSTRUMENT_TYPES = {"EQ", "BE", "BZ", "ETF", "MF", "NCD", "GS", "SG"}

# Canonical NSE ticker overrides for securities with multi-word script names.
# Key: concatenated uppercase script tokens (after stripping instrument suffix).
# Value: correct NSE symbol.
_SYMBOL_OVERRIDES: dict[str, str] = {
    # Previously known aliases
    "ATHERENERGYLIMITED": "ATHERENERG",
    "INDUSTOWERSLIMITED": "INDUSTOWER",
    "PAYTM": "PAYTM",
    "GROWWAMC-GROWWDEFNC": "GROWWDEFNC",
    # Full company names → NSE tickers (backoffice uses company names, not tickers)
    "ADANITOTALGASLIMITED": "ATGL",
    "AMARARAJAENERGY&MOBILITYL": "ARE&M",
    "ANGELONELIMITED": "ANGELONE",
    "BELRISEINDUSTRIESLIMITED": "BELRISE",
    "COMPUTERAGEMANAGEMENTSERVICESLI": "CAMS",
    "DATAPATTERNS(INDIA)LIMITED": "DATAPATTNS",
    "GMRAIRPORTSLIMITED": "GMRAIRPORT",
    "JUBILANTPHARMOVALIMITED": "JUBLPHARMA",
    "JUPITERWAGONSLIMITED": "JWL",
    "MANKINDPHARMALIMITED": "MANKIND",
    "POONAWALLAFINCORPLIMITED": "POONAWALLA",
    "PRIVISPECIALITYCHEMICALSLIM": "PRIVISCL",
    "RAILTELCORPORATIONOFINDIAL": "RAILTEL",
    "SAMVARDHANAMOTHERSONINTERNAT": "MOTHERSON",
    # Delisted / merged → successor tickers
    "ADANITRANS": "ADANIENSOL",       # Adani Transmission merged into Adani Energy Solutions
    "LTI": "LTIM",                    # LTI merged with Mindtree → LTI Mindtree
    "PVR": "PVRINOX",                 # PVR merged with INOX → PVR INOX
    "GMRINFRA": "GMRAIRPORT",         # GMR Infrastructure restructured into GMR Airports
    "AMARAJABAT": "ARE&M",            # Amara Raja Batteries renamed to Amara Raja Energy & Mobility
    "MINDAIND": "UNOMINDA",           # Minda Industries renamed to Uno Minda
    "SUVENPHAR": "SUVEN",             # Suven Pharmaceuticals — NSE ticker is SUVEN
    # Corporate name changes (effective 2026-04-16)
    "TATAMOTORS": "TMPV",             # Tata Motors → Tata Motors Passenger Vehicles (TMPV)
    "ZOMATO": "ETERNAL",              # Zomato → Eternal Limited (ETERNAL)
    "ZOMATOLIMITED": "ETERNAL",       # Full company-name form → Eternal Limited
    "HIL": "BIRLANU",                 # HIL Limited → BirlaNu Limited (BIRLANU)
    "SWANENERGY": "SWANCORP",         # Swan Energy → Swan Corp Limited (SWANCORP)
    # ETFs recorded under AMC product name instead of NSE ticker
    "ICICIPRUDENTIALBSESENSEXET": "SENSEXIETF",
    "MIRAESMALLCAP": "MASMC250",   # Mirae Asset Nifty SmallCap 250 ETF (13 chars → bypasses 12-char ISIN cache regex)
}

# Sector mapping for ETF/commodity instruments
SYMBOL_SECTOR_MAP: dict[str, str] = {
    "GOLDBEES": "Metals",
    "GOLDCASE": "Metals",
    "GOLDSHARE": "Metals",
    "GOLDETF": "Metals",
    "SILVERBEES": "Metals",
    "SILVERETF": "Metals",
    "SILVERSHARE": "Metals",
    "GOLDHERA": "Metals",
    "GOLDPETAL": "Metals",
    "BANKBEES": "Banking",
    "PSUBNKBEES": "Banking",
    "NIFTYBEES": "Diversified",
    "JUNIORBEES": "Diversified",
    "CPSEETF": "Diversified",
    "PHARMABEES": "Pharma",
    "FMCGIETF": "FMCG",
    "HNGSNGBEES": "Diversified",
}

# ── Column layout descriptor ─────────────────────────────────────────────────

TxnColMap = namedtuple("TxnColMap", [
    "has_isin",
    "col_isin",       # None for old format
    "col_script",
    "col_exch",
    "col_stno",
    "col_buy_qty",    # start of Buy block (+0)
    "col_sale_qty",   # start of Sale block (+0)
])


def _detect_col_map(header_row: tuple) -> TxnColMap:
    """Detect column layout from the first header row.

    The header row looks like one of:
      Old:  ('UCC', 'Script', 'Exch', 'Stno', 'Buy', 'Buy', ...)   — 20 cols
      New:  ('UCC', 'ISIN', 'Script', 'Exch', 'Stno', 'Buy', ...)  — 21 cols

    Detection is purely name-based: if col 1 contains 'isin' (case-insensitive)
    the new format is assumed, shifting all subsequent indices by 1.
    """
    has_isin = False
    col_isin: int | None = None

    if len(header_row) > 1 and header_row[1] is not None:
        cell1 = str(header_row[1]).strip().lower()
        if cell1 == "isin":
            has_isin = True
            col_isin = 1
            logger.info("TXN parser: detected 21-col format with ISIN at col 1")
        else:
            logger.info("TXN parser: detected 20-col format (no ISIN column)")

    offset = 1 if has_isin else 0

    # Script, Exch, Stno follow immediately after UCC (+ ISIN if present)
    col_script = 1 + offset
    col_exch = 2 + offset
    col_stno = 3 + offset

    # Buy block starts right after Stno (8 cols wide)
    col_buy_qty = 4 + offset

    # Sale block starts 8 cols after Buy block
    col_sale_qty = col_buy_qty + 8

    return TxnColMap(
        has_isin=has_isin,
        col_isin=col_isin,
        col_script=col_script,
        col_exch=col_exch,
        col_stno=col_stno,
        col_buy_qty=col_buy_qty,
        col_sale_qty=col_sale_qty,
    )


# ── Helper functions ─────────────────────────────────────────────────────────


def classify_sector(symbol: str) -> str:
    """Determine sector from symbol name.

    - Any symbol containing 'liquid' (case-insensitive) → 'Cash'
    - Known ETFs/commodities → mapped sector
    - Everything else → empty string (filled during holdings ingestion from DB)
    """
    if "LIQUID" in symbol.upper():
        return "Cash"
    return SYMBOL_SECTOR_MAP.get(symbol, "")


def _safe_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, returning Decimal('0') for non-numeric."""
    if value is None:
        return Decimal("0")
    try:
        d = Decimal(str(value))
        if d.is_nan():
            return Decimal("0")
        return d
    except (InvalidOperation, ValueError):
        return Decimal("0")


def parse_script(script_raw: str) -> tuple[str, str]:
    """
    Parse script field like "RELIANCE     EQ" into (symbol, instrument_type).

    Handles multi-word names: "Mirae Smallcap ETF" → ("MIRAESMALLCAP", "ETF")
    Single-word: "RELIANCE     EQ" → ("RELIANCE", "EQ")

    Returns:
        Tuple of (symbol, instrument_type). Defaults instrument to "EQ".
    """
    parts = script_raw.strip().split()
    if not parts:
        return script_raw.strip().upper(), "EQ"

    # Last token is instrument type if it's a known type
    if len(parts) >= 2 and parts[-1].upper() in _KNOWN_INSTRUMENT_TYPES:
        instrument = parts[-1].upper()
        name_parts = parts[:-1]
    else:
        instrument = "EQ"
        name_parts = parts

    # For single-word symbols (99% of cases): "RELIANCE" → "RELIANCE"
    # For multi-word: "Mirae Smallcap" → "MIRAESMALLCAP" (concat, no spaces)
    symbol = "".join(p.strip() for p in name_parts).upper()
    # Apply canonical NSE ticker override if known
    symbol = _SYMBOL_OVERRIDES.get(symbol, symbol)
    return symbol, instrument


def _determine_txn_type_buy(stno: str) -> str:
    """Determine transaction type for buy-side entry."""
    if stno.upper() == "BONUS":
        return "BONUS"
    return "BUY"


def _determine_txn_type_sell(stno: str) -> str:
    """Determine transaction type for sell-side entry."""
    if stno.lower().strip() == "corpus":
        return "CORPUS_IN"
    return "SELL"


# ── Main parser ──────────────────────────────────────────────────────────────


def parse_transaction_file(filepath: str | Path) -> list[dict]:
    """
    Parse a PMS Transaction Report .xlsx file.

    Auto-detects column layout (20-col or 21-col with ISIN) from the header row.
    Uses net rate (not cost rate) for pricing — matches backoffice FIFO calculation.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        List of dicts with keys:
            client_code, client_name, date, txn_type, symbol,
            instrument_type, exchange, settlement_no, quantity,
            price, cost_rate, amount, asset_class, sector, isin
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Transaction file not found: {filepath}")

    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Transaction workbook has no active sheet")

    records: list[dict] = []
    current_client_code: str | None = None
    current_client_name: str | None = None
    current_date: datetime | None = None
    row_count = 0
    header_rows_seen = 0  # first 2 rows are headers
    col_map: TxnColMap | None = None
    clients_seen: set[str] = set()

    for row in ws.iter_rows(values_only=True):
        row_count += 1

        # ── First row: detect column layout ─────────────────────────────────
        if header_rows_seen == 0:
            col_map = _detect_col_map(row)
            header_rows_seen += 1
            continue

        # ── Second row: sub-header with column names — skip ──────────────────
        if header_rows_seen == 1:
            header_rows_seen += 1
            continue

        assert col_map is not None  # always set after row 0

        # Pad row to at least col_sale_qty + 8 columns to avoid index errors
        min_cols = col_map.col_sale_qty + 8
        cells = list(row) + [None] * max(0, min_cols - len(row))

        ucc_raw = cells[0]  # col 0 is always UCC

        # Subtotal / grand total rows — UCC is None
        if ucc_raw is None:
            continue

        ucc = str(ucc_raw).strip()
        if not ucc or ucc.lower() == "nan":
            continue

        # Type 1: Client name header — "FULL NAME [CODE]"
        name_match = _NAME_PATTERN.match(ucc)
        if name_match:
            current_client_name = name_match.group(1).strip()
            current_client_code = name_match.group(2).strip()
            current_date = None
            if current_client_code not in clients_seen:
                clients_seen.add(current_client_code)
                logger.info(
                    "TXN parser: processing client %s (%s) — client #%d",
                    current_client_code,
                    current_client_name,
                    len(clients_seen),
                )
            continue

        # Type 2: Date separator — "     Date :DD/MM/YY"
        date_match = _DATE_PATTERN.search(ucc)
        if date_match:
            try:
                current_date = datetime.strptime(date_match.group(1), "%d/%m/%y")
            except ValueError:
                logger.warning("Unparseable date separator: %r", ucc)
            continue

        # Type 3: Transaction data row
        if current_client_code is None or current_date is None:
            continue
        if ucc.rstrip() != current_client_code:
            continue

        script_raw = cells[col_map.col_script]
        if script_raw is None:
            continue
        script_str = str(script_raw).strip()
        if not script_str or script_str.lower() == "nan":
            continue

        symbol, inst_type = parse_script(script_str)
        stno_raw = cells[col_map.col_stno]
        stno = str(stno_raw).strip() if stno_raw is not None else ""
        exch_raw = cells[col_map.col_exch]
        exchange = str(exch_raw).strip() if exch_raw is not None else ""

        # Extract ISIN (None for old format files)
        isin = ""
        if col_map.has_isin and col_map.col_isin is not None:
            isin_raw = cells[col_map.col_isin]
            isin = str(isin_raw).strip() if isin_raw is not None else ""
            if isin.lower() in ("none", "nan"):
                isin = ""

        asset_class = "CASH" if "LIQUID" in symbol.upper() else "EQUITY"
        sector = classify_sector(symbol)

        # Buy block: col_buy_qty + offsets
        buy_qty = _safe_decimal(cells[col_map.col_buy_qty])
        buy_rate = _safe_decimal(cells[col_map.col_buy_qty + 1])       # net rate
        buy_cost_rate = _safe_decimal(cells[col_map.col_buy_qty + 5])  # all-in cost rate
        buy_amount = _safe_decimal(cells[col_map.col_buy_qty + 6])     # amount with cost

        # Sale block: col_sale_qty + offsets
        sale_qty = _safe_decimal(cells[col_map.col_sale_qty])
        sale_rate = _safe_decimal(cells[col_map.col_sale_qty + 1])       # net rate
        sale_cost_rate = _safe_decimal(cells[col_map.col_sale_qty + 5])  # all-in cost rate
        sale_amount = _safe_decimal(cells[col_map.col_sale_qty + 6])     # amount with cost

        # A single row can have BOTH buy and sell — check independently
        if buy_qty > 0:
            records.append({
                "client_code": current_client_code,
                "client_name": current_client_name,
                "date": current_date,
                "txn_type": _determine_txn_type_buy(stno),
                "symbol": symbol,
                "instrument_type": inst_type,
                "exchange": exchange,
                "settlement_no": stno,
                "quantity": buy_qty,
                "price": buy_rate,
                "cost_rate": buy_cost_rate,
                "amount": buy_amount,
                "asset_class": asset_class,
                "sector": sector,
                "isin": isin,
            })

        if sale_qty > 0:
            records.append({
                "client_code": current_client_code,
                "client_name": current_client_name,
                "date": current_date,
                "txn_type": _determine_txn_type_sell(stno),
                "symbol": symbol,
                "instrument_type": inst_type,
                "exchange": exchange,
                "settlement_no": stno,
                "quantity": sale_qty,
                "price": sale_rate,
                "cost_rate": sale_cost_rate,
                "amount": sale_amount,
                "asset_class": asset_class,
                "sector": sector,
                "isin": isin,
            })

    wb.close()
    logger.info(
        "TXN parser complete: %d records from %d clients (%d rows scanned)",
        len(records),
        len(clients_seen),
        row_count,
    )
    return records
